import argparse
import hashlib
import os
import re
import ipaddress
from datetime import datetime

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_VERSION = "1.0.0-fap"


# -----------------------------
# Helpers
# -----------------------------
ANY_TOKENS = {"any", "all", "*", "any (any)"}

def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def normalize_text(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

def normalize_any_field(x: str) -> str:
    v = normalize_text(x)
    if not v:
        return ""
    if v.lower() in ANY_TOKENS:
        return "ANY"
    return v

def services_contains_any(x: str) -> bool:
    v = normalize_text(x).lower()
    if not v:
        return False
    if v in ANY_TOKENS:
        return True
    # "Any" can appear inside lists or combined strings
    return bool(re.search(r"\bany\b", v))

def normalize_services(x: str) -> str:
    v = normalize_text(x)
    if not v:
        return ""
    return "ANY" if services_contains_any(v) else v

def normalize_logging(x: str) -> str:
    v = normalize_text(x).lower()
    if v in {"log", "logged", "yes", "true", "enabled"}:
        return "Enabled"
    if v in {"", "none", "no", "false", "disabled"}:
        return "None"
    # Track modes (e.g., "Detailed Log", "Account", etc.) count as enabled for gap reporting
    return "Enabled"

def is_overly_broad(value: str, broad_prefix_threshold: int = 24) -> bool:
    """
    Flags ANY or broad subnets.
    Default: anything broader than /24 (i.e., /23, /16, /8) is considered broad.
    """
    v = normalize_text(value)
    if not v:
        return False
    if v == "ANY":
        return True
    try:
        if "/" in v:
            net = ipaddress.ip_network(v, strict=False)
            return net.prefixlen < broad_prefix_threshold
    except ValueError:
        return False
    return False

def unused_flag(hit_value):
    if pd.isna(hit_value):
        return "Unknown"
    try:
        hv = float(hit_value)
    except Exception:
        return "Unknown"
    if hv == 0:
        return "Yes"
    if hv > 0:
        return "No"
    return "Unknown"

def safe_lower(x: str) -> str:
    return normalize_text(x).lower()


# -----------------------------
# Mapping file (STIG/SRG/CIS/NIST)
# -----------------------------
def load_control_mapping(mapping_path: str) -> pd.DataFrame:
    """
    Expected columns in mapping CSV:
      Finding_Key,DoD_Stig_Ref,SRG_Ref,CIS_CP_Benchmark_Ref,NIST_Control
    """
    if not mapping_path:
        return pd.DataFrame(columns=[
            "Finding_Key", "DoD_Stig_Ref", "SRG_Ref", "CIS_CP_Benchmark_Ref", "NIST_Control"
        ])

    if not os.path.isfile(mapping_path):
        raise FileNotFoundError(f"Mapping file not found: {mapping_path}")

    m = pd.read_csv(mapping_path, dtype=str, sep="\t").fillna("")
    required = {"Finding_Key", "DoD_Stig_Ref", "SRG_Ref", "CIS_CP_Benchmark_Ref", "NIST_Control"}
    missing = required - set(m.columns)
    if missing:
        raise ValueError(f"Mapping file missing required columns: {sorted(missing)}")

    m["Finding_Key"] = m["Finding_Key"].astype(str).str.strip()
    return m

def build_map_lookup(map_df: pd.DataFrame) -> dict:
    if map_df.empty:
        return {}
    return map_df.set_index("Finding_Key").to_dict(orient="index")

def concat_refs(keys, lookup, field: str) -> str:
    vals = []
    for k in keys:
        if k in lookup:
            v = normalize_text(lookup[k].get(field, ""))
            if v:
                vals.append(v)
    # de-dupe preserving order
    seen = set()
    out = []
    for v in vals:
        if v not in seen:
            out.append(v)
            seen.add(v)
    return "; ".join(out)


# -----------------------------
# Excel formatting helpers
# -----------------------------
def autosize_columns(ws, max_width: int = 60):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = min(max(10, max_len + 2), max_width)
        ws.column_dimensions[col_letter].width = width

def freeze_and_filter(ws, header_row: int = 1):
    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = ws.dimensions

def add_table_style(ws, table_name: str = "RulesTable"):
    # Add an Excel table if possible (only works well for rectangular ranges)
    try:
        ref = ws.dimensions
        tab = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        # If table creation fails (rare), ignore
        pass


# -----------------------------
# Main logic
# -----------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Standardize SmartConsole CSV export, auto-flag findings, and generate FAP/FAPR-ready Excel."
    )
    parser.add_argument("--input", "-i", required=True, help="Path to SmartConsole CSV export")
    parser.add_argument("--output", "-o", default="", help="Output XLSX path (default: <input>_STANDARDIZED.xlsx)")
    parser.add_argument("--mapping", "-m", default="control_mapping.csv", help="Control mapping CSV (STIG/SRG/CIS/NIST).")
    parser.add_argument("--broad-threshold", type=int, default=24, help="Prefixlen threshold: < value is considered broad (default: 24).")
    parser.add_argument("--standards", default="NIST SP 800-53 Rev. 5; DoDI 8500.01; DoDI 8510.01 (RMF); DoD Firewall STIG (derived from Firewall SRG v3r3)",
                        help="Standards reference string to embed in output.")
    args = parser.parse_args()

    input_file = args.input.strip()
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    base_name = os.path.splitext(input_file)[0]
    output_file = args.output.strip() or f"{base_name}_STANDARDIZED.xlsx"

    # Load mapping
    map_df = load_control_mapping(args.mapping) if args.mapping else pd.DataFrame()
    map_lookup = build_map_lookup(map_df)

    # Load CSV
    df = pd.read_csv(input_file, encoding="utf-8-sig", dtype=str).fillna("")
    df.columns = df.columns.str.strip()

    # Expect your known headers
    expected_cols = [
        "No.", "Type", "Name", "Source", "Destination", "VPN",
        "Services & Applications", "Action", "Track", "Install On"
    ]
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        raise KeyError(
            "CSV headers do not match expected SmartConsole export format.\n"
            f"Missing columns: {missing}\n"
            f"Detected columns: {list(df.columns)}"
        )

    # Rename
    df = df.rename(columns={
        "No.": "Rule_Number",
        "Type": "Rule_Type",
        "Name": "Rule_Name",
        "Source": "Source",
        "Destination": "Destination",
        "VPN": "VPN",
        "Services & Applications": "Services",
        "Action": "Action",
        "Track": "Logging",
        "Install On": "Install_On"
    })

    # Keep the core columns
    df = df[[
        "Rule_Number", "Rule_Type", "Rule_Name", "Source", "Destination",
        "VPN", "Services", "Action", "Logging", "Install_On"
    ]].copy()

    # Normalize
    df["Rule_Number"] = pd.to_numeric(df["Rule_Number"], errors="coerce")
    df["Rule_Type"] = df["Rule_Type"].apply(normalize_text)
    df["Rule_Name"] = df["Rule_Name"].apply(normalize_text)

    for col in ["Source", "Destination"]:
        df[col] = df[col].apply(normalize_any_field)

    df["Services"] = df["Services"].apply(normalize_services)
    df["Action"] = df["Action"].apply(normalize_text)
    df["Logging"] = df["Logging"].apply(normalize_logging)
    df["VPN"] = df["VPN"].apply(normalize_text)
    df["Install_On"] = df["Install_On"].apply(normalize_text)

    # Hit count: not present in your export => Unknown
    df["Hit_Count"] = pd.NA
    df["Last_Hit_Date"] = "Unknown"
    df["Unused_Rule"] = "Unknown"

    # Filter enforceable rules:
    # SmartConsole exports vary, so we do a SAFE filter:
    # - Keep rows where Rule_Number is a number AND Action is non-empty
    # This avoids scoring section headers / separators that often have no action.
    df = df[df["Rule_Number"].notna()].copy()
    df = df[df["Action"].astype(str).str.strip() != ""].copy()

    # Flags
    df["Flag_Any_SrcDst"] = df.apply(lambda r: "Yes" if r["Source"] == "ANY" or r["Destination"] == "ANY" else "No", axis=1)
    df["Flag_Broad_SrcDst"] = df.apply(lambda r: "Yes" if is_overly_broad(r["Source"], args.broad_threshold) or is_overly_broad(r["Destination"], args.broad_threshold) else "No", axis=1)
    df["Flag_Any_Services"] = df["Services"].apply(lambda x: "Yes" if x == "ANY" else "No")
    df["Flag_No_Logging"] = df.apply(lambda r: "Yes" if safe_lower(r["Action"]) == "accept" and r["Logging"] == "None" else "No", axis=1)
    df["VPN_Constrained"] = df["VPN"].apply(lambda x: "Yes" if normalize_text(x) not in {"", "Any", "ANY", "None", "N/A"} else "No")

    # Risk assessment (explainable)
    def assess_risk(row):
        action = safe_lower(row["Action"])
        if action != "accept":
            return "Low"

        bad = 0
        if row["Flag_Any_SrcDst"] == "Yes":
            bad += 1
        if row["Flag_Broad_SrcDst"] == "Yes":
            bad += 1
        if row["Flag_Any_Services"] == "Yes":
            bad += 1
        if row["Flag_No_Logging"] == "Yes":
            bad += 1

        if bad >= 2:
            return "High"
        if bad == 1:
            return "Medium"
        return "Low"

    df["Risk_Level"] = df.apply(assess_risk, axis=1)

    # Evidence summary (auditor-friendly)
    def evidence(row):
        parts = [
            f"Action={normalize_text(row['Action'])}",
            f"Source={normalize_text(row['Source'])}",
            f"Destination={normalize_text(row['Destination'])}",
            f"Services={normalize_text(row['Services'])}",
            f"Logging={normalize_text(row['Logging'])}",
            f"VPN={normalize_text(row['VPN']) or 'N/A'}"
        ]
        return "; ".join(parts)

    df["Evidence_Summary"] = df.apply(evidence, axis=1)

    # Finding keys (used to map STIG/SRG/CIS/NIST)
    def build_finding_keys(row):
        keys = []
        if safe_lower(row["Action"]) != "accept":
            return keys

        if row["Flag_Any_SrcDst"] == "Yes":
            keys.append("ANY_SOURCE_DEST")
        elif row["Flag_Broad_SrcDst"] == "Yes":
            keys.append("BROAD_SUBNET")

        if row["Flag_Any_Services"] == "Yes":
            keys.append("ANY_SERVICES")

        if row["Flag_No_Logging"] == "Yes":
            keys.append("NO_LOGGING")

        # Only include UNUSED_RULE if known; your export is unknown by design
        if row.get("Unused_Rule", "Unknown") == "Yes":
            keys.append("UNUSED_RULE")

        return keys

    df["Finding_Keys"] = df.apply(build_finding_keys, axis=1)

    # Human-readable finding text + remediation (based on flags)
    def generate_finding(row):
        if safe_lower(row["Action"]) != "accept":
            return ""
        findings = []
        if row["Flag_Any_SrcDst"] == "Yes":
            findings.append("Overly broad source/destination (ANY), violates least-privilege")
        elif row["Flag_Broad_SrcDst"] == "Yes":
            findings.append("Overly broad source/destination (large subnet), violates least-privilege")
        if row["Flag_Any_Services"] == "Yes":
            findings.append("Unrestricted services (ANY), violates least-privilege")
        if row["Flag_No_Logging"] == "Yes":
            findings.append("Traffic allowed without logging")
        # Unused rule not asserted unless hit data exists
        return "; ".join(findings)

    def generate_remediation(row):
        if not row["Finding"]:
            return ""
        recs = []
        f = row["Finding"].lower()
        if "source/destination" in f:
            recs.append("Restrict source/destination to operationally required scope")
        if "services" in f:
            recs.append("Restrict services to required ports/protocols only")
        if "without logging" in f:
            recs.append("Enable logging (or document/approve exception)")
        # VPN context note
        if row.get("VPN_Constrained") == "Yes":
            recs.append("Verify exposure is constrained by VPN community and document boundary assumptions")
        return "; ".join(recs)

    df["Finding"] = df.apply(generate_finding, axis=1)
    df["Remediation"] = df.apply(generate_remediation, axis=1)

    # Map compliance references using mapping CSV
    df["DoD_Stig_Ref"] = df["Finding_Keys"].apply(lambda keys: concat_refs(keys, map_lookup, "DoD_Stig_Ref"))
    df["SRG_Ref"] = df["Finding_Keys"].apply(lambda keys: concat_refs(keys, map_lookup, "SRG_Ref"))
    df["CIS_CP_Benchmark_Ref"] = df["Finding_Keys"].apply(lambda keys: concat_refs(keys, map_lookup, "CIS_CP_Benchmark_Ref"))
    df["NIST_Control"] = df["Finding_Keys"].apply(lambda keys: concat_refs(keys, map_lookup, "NIST_Control"))

    # Analyst workflow columns (for defensible FAPR process)
    df["Auto_Flag"] = df["Risk_Level"].apply(lambda x: "Yes" if x in {"High", "Medium"} else "No")
    df["Analyst_Status"] = "Unreviewed"  # Confirmed / False Positive / Accepted Risk / Needs More Info
    df["Risk_Final"] = ""                # allow analyst override
    df["Analyst_Notes"] = ""

    # Standards reference column
    df["Compliance_Standards_Referenced"] = args.standards

    # Pivot summaries
    pivot_risk = pd.pivot_table(df, index="Risk_Level", values="Rule_Number", aggfunc="count").rename(columns={"Rule_Number": "Rule_Count"})
    pivot_action = pd.pivot_table(df, index="Action", values="Rule_Number", aggfunc="count").rename(columns={"Rule_Number": "Rule_Count"})
    pivot_any = pd.pivot_table(df, index="Flag_Any_SrcDst", values="Rule_Number", aggfunc="count").rename(columns={"Rule_Number": "Rule_Count"})
    pivot_logging = pd.pivot_table(df, index="Flag_No_Logging", values="Rule_Number", aggfunc="count").rename(columns={"Rule_Number": "Rule_Count"})
    pivot_vpn = pd.pivot_table(df, index="VPN_Constrained", values="Rule_Number", aggfunc="count").rename(columns={"Rule_Number": "Rule_Count"})

    # Run metadata
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    input_hash = sha256_file(input_file)
    metadata = pd.DataFrame([
        {"Key": "Script_Version", "Value": SCRIPT_VERSION},
        {"Key": "Run_Timestamp_Local", "Value": run_time},
        {"Key": "Input_File", "Value": os.path.abspath(input_file)},
        {"Key": "Input_SHA256", "Value": input_hash},
        {"Key": "Broad_Subnet_Threshold", "Value": f"< /{args.broad_threshold} is broad"},
        {"Key": "Standards_Referenced", "Value": args.standards},
        {"Key": "Mapping_File", "Value": os.path.abspath(args.mapping) if args.mapping else ""},
    ])

    # Output columns (final ordering)
    ordered_cols = [
        "Rule_Number", "Rule_Type", "Rule_Name",
        "Source", "Destination", "VPN", "Services",
        "Action", "Logging", "Install_On",
        "Risk_Level", "Auto_Flag",
        "Finding", "Remediation", "Evidence_Summary",
        "NIST_Control", "DoD_Stig_Ref", "SRG_Ref", "CIS_CP_Benchmark_Ref",
        "Analyst_Status", "Risk_Final", "Analyst_Notes",
        "Hit_Count", "Last_Hit_Date", "Unused_Rule",
        "Flag_Any_SrcDst", "Flag_Broad_SrcDst", "Flag_Any_Services", "Flag_No_Logging", "VPN_Constrained",
        "Compliance_Standards_Referenced"
    ]
    df = df[[c for c in ordered_cols if c in df.columns]].copy()

    # Write Excel
    if os.path.exists(output_file):
        os.remove(output_file)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Firewall Rules & Findings", index=False)
        pivot_risk.to_excel(writer, sheet_name="Summary - Risk Levels")
        pivot_action.to_excel(writer, sheet_name="Summary - Actions")
        pivot_any.to_excel(writer, sheet_name="Summary - ANY Usage")
        pivot_logging.to_excel(writer, sheet_name="Summary - Logging Gaps")
        pivot_vpn.to_excel(writer, sheet_name="Summary - VPN Constrained")
        metadata.to_excel(writer, sheet_name="Run Metadata", index=False)
        if not map_df.empty:
            map_df.to_excel(writer, sheet_name="Mappings - STIG SRG CIS", index=False)

        # Excel formatting (openpyxl)
        wb = writer.book

        ws = wb["Firewall Rules & Findings"]
        freeze_and_filter(ws)
        autosize_columns(ws)
        add_table_style(ws, table_name="RulesTable")

        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Format summary sheets
        for sheet in [
            "Summary - Risk Levels", "Summary - Actions", "Summary - ANY Usage",
            "Summary - Logging Gaps", "Summary - VPN Constrained", "Run Metadata"
        ]:
            if sheet in wb.sheetnames:
                w = wb[sheet]
                freeze_and_filter(w)
                autosize_columns(w)
                for cell in w[1]:
                    cell.font = Font(bold=True)

        if "Mappings - STIG SRG CIS" in wb.sheetnames:
            w = wb["Mappings - STIG SRG CIS"]
            freeze_and_filter(w)
            autosize_columns(w)
            for cell in w[1]:
                cell.font = Font(bold=True)

    print(f"✔ Consolidated rules and findings exported to: {output_file}")
    print(f"✔ Input SHA256: {input_hash}")
    if args.mapping:
        print(f"✔ Mapping file used: {args.mapping}")


if __name__ == "__main__":
    main()

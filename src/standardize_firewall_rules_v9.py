import pandas as pd
import os
import ipaddress
import json
import re
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

"""
standardize_firewall_rules_v8.py

Fixes the "blank STIG columns" issue by adding:
- Auto-matching of STIG Vuln IDs from firewall_security_requirements_guide.json (keyword scoring)
- Keeps manual overrides via FINDING_TO_STIG_VIDS (authoritative, audit-friendly)
- Adds match metadata columns:
  - STIG_Match_Method (Manual/Auto/None)
  - STIG_Match_Score (auto-score)
  - STIG_Match_Notes (e.g., "Top 2 matches by keyword score")

Still includes:
- CIS mapping from cis_checkpoint_benchmark_v1_1_0_mapping.json (generated from CIS PDF)
- Risk color coding and STIG CAT color coding
- Metrics Summary sheet
"""

# -----------------------------
# Paths (adjust if needed)
# -----------------------------
STIG_JSON_PATH = "firewall_security_requirements_guide.json"
CIS_MAPPING_JSON_PATH = "cis_checkpoint_benchmark_v1_1_0_mapping.json"

# -----------------------------
# Load STIG/SRG guide
# -----------------------------
STIG_RULES_BY_VID = {}
STIG_SEARCH_INDEX = []  # list of tuples (vid, severity, title, discussion, haystack)
if os.path.exists(STIG_JSON_PATH):
    with open(STIG_JSON_PATH, "r", encoding="utf-8") as f:
        stig = json.load(f)

    for g in stig.get("groups", []):
        vid = str(g.get("groupId", "")).strip()
        if not vid:
            continue
        STIG_RULES_BY_VID[vid] = g

        sev = str(g.get("ruleSeverity", "")).strip()
        title = str(g.get("ruleTitle", "")).strip()
        disc = str(g.get("ruleVulnDiscussion", "")).strip()
        hay = (title + " " + disc).lower()
        STIG_SEARCH_INDEX.append((vid, sev, title, disc, hay))

def severity_to_cat(sev: str) -> str:
    sev = (sev or "").strip().lower()
    if sev == "high":
        return "CAT I"
    if sev == "medium":
        return "CAT II"
    if sev == "low":
        return "CAT III"
    return ""

def pick_highest_severity(severities):
    order = {"high": 3, "medium": 2, "low": 1}
    best = ""
    best_score = 0
    for s in severities:
        k = (s or "").strip().lower()
        if order.get(k, 0) > best_score:
            best_score = order[k]
            best = k
    return best

# -----------------------------
# STIG mapping: Manual overrides (authoritative)
# Populate these once you finalize your environment’s mapping.
# -----------------------------
FINDING_TO_STIG_VIDS = {
    # Mapped using Firewall SRG entries from STIG Viewer (Firewall Security Requirements Guide, 2024-12-04):
    # - V-206694 explicitly calls out avoiding rules containing "any" for host/destination/protocol/port and requires deny-by-default/permit-by-exception.
    # - V-206674 requires filtering using src/dst IPs and ports (packet headers/attributes).
    # - V-206695 ties permitted traffic to explicit ports/protocols aligned to PPSM CAL and VAs.
    # - V-206699 requires sending traffic logs to a central audit server.
    # - V-206681 requires traffic logs capture event source info (e.g., source IP).
    # - V-206682 requires traffic logs capture outcome (success/failure of rule application).

    # Broad/ANY scoping findings
    "Overly broad source/destination (ANY)": ["V-206694", "V-206674"],
    "Overly broad source/destination (large subnet)": ["V-206694", "V-206674"],
    "Unrestricted services (ANY)": ["V-206694", "V-206695"],

    # Logging findings (when your rulebase allows traffic but logging is off / insufficient)
    "Traffic allowed without logging": ["V-206699", "V-206681", "V-206682"],

    # These are policy-hygiene/optimization findings; the Firewall SRG doesn't map cleanly 1:1 to them.
    # Keep blank until you decide which STIG/STIG-implementation (e.g., Check Point STIG) to bind them to.
    # "Unused rule (0 hits)": [],
    # "Shadowed rule (covered by earlier rule)": [],
    # "Redundant rule (duplicate match conditions)": [],
}


# Auto-match keywords per finding part (tune as you like)
FINDING_KEYWORDS = {
    "Overly broad source/destination (ANY)": ["any", "source", "destination", "restrict", "scope", "least privilege"],
    "Overly broad source/destination (large subnet)": ["subnet", "network", "restrict", "scope", "least privilege"],
    "Unrestricted services (ANY)": ["service", "services", "port", "protocol", "any", "restrict"],
    "Traffic allowed without logging": ["log", "logging", "audit", "track", "records"],
    "Unused rule (0 hits)": ["hit", "count", "unused", "review", "remove"],
    "Shadowed rule (covered by earlier rule)": ["shadow", "rulebase", "optimiz", "review", "cleanup"],
    "Redundant rule (duplicate match conditions)": ["redundant", "rulebase", "optimiz", "review", "cleanup"],
}

def keyword_score(haystack: str, keywords):
    # simple scoring: +2 if keyword appears in title-ish terms; +1 otherwise
    score = 0
    for kw in keywords:
        k = kw.lower().strip()
        if not k:
            continue
        if k in haystack:
            score += 1
    return score

def auto_match_stig_ids(finding_part: str, top_n=2):
    """
    Returns (vids, score, notes)
    - vids: list[str] of top matches by keyword score
    - score: best score
    - notes: short text
    """
    keywords = FINDING_KEYWORDS.get(finding_part, [])
    if not keywords or not STIG_SEARCH_INDEX:
        return [], 0, "No keywords/index"
    scored = []
    for vid, sev, title, disc, hay in STIG_SEARCH_INDEX:
        s = keyword_score(hay, keywords)
        if s > 0:
            scored.append((s, vid, title))
    scored.sort(reverse=True, key=lambda x: x[0])
    top = scored[:top_n]
    vids = [v for _, v, _ in top]
    best = top[0][0] if top else 0
    notes = f"Top {len(vids)} by keyword score" if vids else "No match"
    return vids, best, notes

# -----------------------------
# Load CIS mapping (generated from the PDF)
# -----------------------------
CIS_FINDING_TO_CONTROLS = {}
if os.path.exists(CIS_MAPPING_JSON_PATH):
    with open(CIS_MAPPING_JSON_PATH, "r", encoding="utf-8") as f:
        cis_map = json.load(f)
    CIS_FINDING_TO_CONTROLS = cis_map.get("finding_to_controls", {})

def format_cis_refs(finding_parts):
    refs = []
    for part in finding_parts or []:
        for c in CIS_FINDING_TO_CONTROLS.get(part, []):
            cid = str(c.get("id", "")).strip()
            title = str(c.get("title", "")).strip()
            if cid and title:
                refs.append(f"{cid} - {title}")
            elif cid:
                refs.append(cid)
    out, seen = [], set()
    for r in refs:
        if r not in seen:
            out.append(r)
            seen.add(r)
    return "; ".join(out)

# -----------------------------
# Input CSV
# -----------------------------
input_file = input("Enter SmartConsole CSV filename (e.g., FAP Lab.csv): ").strip()
if not os.path.isfile(input_file):
    raise FileNotFoundError(f"File not found: {input_file}")

base_name = os.path.splitext(input_file)[0]
output_file = f"{base_name}_STANDARDIZED.xlsx"

df = pd.read_csv(input_file, encoding="utf-8-sig", dtype=str).fillna("")
df.columns = df.columns.str.strip()

def find_col(possible_names):
    for name in possible_names:
        if name in df.columns:
            return name
    return None

col_no = find_col(["No.", "Rule No.", "Rule Number", "Rule #"])
col_name = find_col(["Name", "Rule Name"])
col_src = find_col(["Source", "Src"])
col_dst = find_col(["Destination", "Dest", "Dst"])
col_svc = find_col(["Services & Applications", "Services", "Service"])
col_action = find_col(["Action"])
col_track = find_col(["Track", "Logging", "Log"])
col_install = find_col(["Install On", "Install On (Policy Targets)", "Policy Targets"])

required = [col_no, col_name, col_src, col_dst, col_svc, col_action]
if any(c is None for c in required):
    raise KeyError(f"CSV is missing required columns. Detected columns: {list(df.columns)}")

df = df.rename(columns={
    col_no: "Rule_Number",
    col_name: "Rule_Name",
    col_src: "Source",
    col_dst: "Destination",
    col_svc: "Services",
    col_action: "Action",
    (col_track or "Track"): "Logging",
    (col_install or "Install On"): "Install_On",
})

for opt in ["Logging", "Install_On"]:
    if opt not in df.columns:
        df[opt] = ""

df = df[["Rule_Number", "Rule_Name", "Source", "Destination", "Services", "Action", "Logging", "Install_On"]]

# -----------------------------
# Normalize values
# -----------------------------
ANY_TOKENS = {"any", "all", "*", "any (any)"}

def normalize_any(value: str) -> str:
    v = str(value).strip()
    if not v:
        return ""
    return "ANY" if v.lower() in ANY_TOKENS else v

def contains_any_token(value: str) -> bool:
    v = str(value).lower()
    return v.strip() in ANY_TOKENS or "any" in v

for col in ["Source", "Destination"]:
    df[col] = df[col].apply(normalize_any)

df["Services"] = df["Services"].apply(lambda x: "ANY" if contains_any_token(x) else str(x).strip())

def normalize_logging(x: str) -> str:
    v = str(x).strip().lower()
    if v in ["log", "logged", "enabled", "true", "yes"]:
        return "Enabled"
    if v in ["", "none", "no", "disabled", "false"]:
        return "None"
    return "Enabled"

df["Logging"] = df["Logging"].apply(normalize_logging)
df["Rule_Number"] = pd.to_numeric(df["Rule_Number"], errors="coerce")

# -----------------------------
# Hit-count (tri-state)
# -----------------------------
hit_col = find_col(["Hit Count", "Hits", "Hit_Count"])
last_hit_col = find_col(["Last Hit", "Last Hit Date", "Last_Hit_Date"])

df["Hit_Count"] = pd.to_numeric(df.get(hit_col, ""), errors="coerce") if hit_col else pd.NA
df["Last_Hit_Date"] = df.get(last_hit_col, "").replace({"": "Unknown"}) if last_hit_col else "Unknown"

def unused_flag(hit):
    if pd.isna(hit):
        return "Unknown"
    return "Yes" if hit == 0 else "No"

df["Unused_Rule"] = df["Hit_Count"].apply(unused_flag)

# -----------------------------
# Least-privilege detection
# -----------------------------
def is_overly_broad(value: str) -> bool:
    if value == "ANY":
        return True
    v = str(value).strip()
    try:
        if "/" in v:
            net = ipaddress.ip_network(v, strict=False)
            return net.prefixlen < 24
    except ValueError:
        return False
    return False

df["Flag_Any_SrcDst"] = df.apply(lambda r: "Yes" if r["Source"] == "ANY" or r["Destination"] == "ANY" else "No", axis=1)
df["Flag_Any_Services"] = df["Services"].apply(lambda x: "Yes" if x == "ANY" else "No")
df["Flag_Broad_SrcDst"] = df.apply(lambda r: "Yes" if is_overly_broad(r["Source"]) or is_overly_broad(r["Destination"]) else "No", axis=1)
df["Flag_No_Logging"] = df.apply(lambda r: "Yes" if str(r["Action"]).strip().lower() == "accept" and r["Logging"] == "None" else "No", axis=1)

# -----------------------------
# Shadowed / Redundant detection (best-effort)
# -----------------------------
def _parse_network(v: str):
    v = str(v).strip()
    if not v or v == "ANY":
        return None
    try:
        if "/" in v:
            return ipaddress.ip_network(v, strict=False)
        return ipaddress.ip_network(v + "/32", strict=False)
    except ValueError:
        return None

def _covers(a: str, b: str) -> bool:
    a = str(a).strip()
    b = str(b).strip()
    if a == "ANY":
        return True
    if b == "ANY":
        return False
    na = _parse_network(a)
    nb = _parse_network(b)
    return bool(na and nb and nb.subnet_of(na))

def _svc_covers(a: str, b: str) -> bool:
    a = str(a).strip()
    b = str(b).strip()
    if a == "ANY":
        return True
    return a == b and a != ""

df = df.sort_values(by=["Rule_Number"], na_position="last").reset_index(drop=True)

dedup_key_cols = ["Source", "Destination", "Services", "Action", "Install_On"]
df["Redundant_Key"] = df[dedup_key_cols].astype(str).agg("|".join, axis=1)
df["Flag_Redundant"] = df.duplicated(subset=["Redundant_Key"], keep="first").map({True: "Yes", False: "No"})

shadowed = []
shadowed_by = []
for i, row in df.iterrows():
    is_shadow = "No"
    by_rule = ""
    if str(row["Action"]).strip().lower() == "accept":
        for j in range(0, i):
            prev = df.loc[j]
            if str(prev["Action"]).strip().lower() != "accept":
                continue
            if str(prev["Install_On"]).strip() and str(row["Install_On"]).strip():
                if str(prev["Install_On"]).strip() != str(row["Install_On"]).strip():
                    continue
            if _covers(prev["Source"], row["Source"]) and _covers(prev["Destination"], row["Destination"]) and _svc_covers(prev["Services"], row["Services"]):
                is_shadow = "Yes"
                by_rule = str(prev.get("Rule_Number", "")).strip()
                break
    shadowed.append(is_shadow)
    shadowed_by.append(by_rule)

df["Flag_Shadowed"] = shadowed
df["Shadowed_By_Rule"] = shadowed_by

# -----------------------------
# Risk assessment (shadowed/redundant = hygiene, not risk)
# -----------------------------
def assess_risk(row):
    action = str(row["Action"]).strip().lower()
    if action != "accept":
        return "Low"
    bad = sum([
        row["Flag_Any_SrcDst"] == "Yes",
        row["Flag_Broad_SrcDst"] == "Yes",
        row["Flag_Any_Services"] == "Yes",
        row["Flag_No_Logging"] == "Yes",
    ])
    if bad >= 2:
        return "High"
    if bad == 1:
        return "Medium"
    return "Low"

df["Risk_Level"] = df.apply(assess_risk, axis=1)

# -----------------------------
# Findings (multi-part)
# -----------------------------
def generate_finding_parts(row):
    action = str(row["Action"]).strip().lower()
    parts = []
    if action == "accept":
        if row["Flag_Any_SrcDst"] == "Yes":
            parts.append("Overly broad source/destination (ANY)")
        elif row["Flag_Broad_SrcDst"] == "Yes":
            parts.append("Overly broad source/destination (large subnet)")
        if row["Flag_Any_Services"] == "Yes":
            parts.append("Unrestricted services (ANY)")
        if row["Flag_No_Logging"] == "Yes":
            parts.append("Traffic allowed without logging")
        if row["Unused_Rule"] == "Yes":
            parts.append("Unused rule (0 hits)")
    if row.get("Flag_Shadowed") == "Yes":
        parts.append("Shadowed rule (covered by earlier rule)")
    if row.get("Flag_Redundant") == "Yes":
        parts.append("Redundant rule (duplicate match conditions)")
    return parts

df["Finding_Parts"] = df.apply(generate_finding_parts, axis=1)
df["Finding"] = df["Finding_Parts"].apply(lambda p: "; ".join(p) if p else "")

# -----------------------------
# Remediation
# -----------------------------
def generate_remediation(parts):
    if not parts:
        return ""
    joined = " ".join(parts).lower()
    recs = []
    if "source/destination" in joined:
        recs.append("Restrict source/destination to operationally required scope")
    if "services" in joined:
        recs.append("Restrict services to required ports/protocols only")
    if "without logging" in joined:
        recs.append("Enable logging (or justify and document exception)")
    if "unused rule" in joined:
        recs.append("Remove unused rule or validate/justify operational requirement")
    if "shadowed rule" in joined:
        recs.append("Remove shadowed rule or document precedence/exception rationale")
    if "redundant rule" in joined:
        recs.append("Consolidate redundant rules to reduce complexity and error risk")
    out, seen = [], set()
    for r in recs:
        if r not in seen:
            out.append(r); seen.add(r)
    return "; ".join(out)

df["Remediation"] = df["Finding_Parts"].apply(generate_remediation)

# -----------------------------
# NIST mapping (simple)
# -----------------------------
def map_nist(parts):
    if not parts:
        return ""
    controls = set()
    joined = " ".join(parts).lower()
    if "source/destination" in joined or "unrestricted services" in joined:
        controls.update(["AC-6", "SC-7"])
    if "without logging" in joined:
        controls.add("AU-6")
    if "unused rule" in joined or "shadowed rule" in joined or "redundant rule" in joined:
        controls.add("CM-6")
    return ", ".join(sorted(controls))

df["NIST_Control"] = df["Finding_Parts"].apply(map_nist)

# -----------------------------
# STIG mapping: Vuln IDs + Titles + Severity + CAT (+ match metadata)
# -----------------------------
def map_stig(parts):
    vids = []
    severities = []
    titles = []
    method = "None"
    best_score = 0
    notes = ""

    for part in parts or []:
        # Manual override first
        mapped = FINDING_TO_STIG_VIDS.get(part, [])
        if mapped:
            method = "Manual"
            for vid in mapped:
                if vid and vid not in vids:
                    vids.append(vid)
        else:
            # Auto-match fallback
            auto_vids, score, auto_notes = auto_match_stig_ids(part, top_n=2)
            if auto_vids:
                method = "Auto" if method != "Manual" else method
                best_score = max(best_score, score)
                notes = auto_notes
                for vid in auto_vids:
                    if vid and vid not in vids:
                        vids.append(vid)

    # Enrich from guide
    for vid in vids:
        rule = STIG_RULES_BY_VID.get(vid, {})
        sev = str(rule.get("ruleSeverity", "")).lower().strip()
        if sev:
            severities.append(sev)
        title = str(rule.get("ruleTitle", "")).strip()
        if title:
            titles.append(f"{vid}: {title}")

    top_sev = pick_highest_severity(severities)
    cat = severity_to_cat(top_sev)

    return (
        "; ".join(vids),
        top_sev.title() if top_sev else "",
        cat,
        " | ".join(titles),
        method,
        best_score,
        notes
    )

stig_cols = df["Finding_Parts"].apply(map_stig)
df["STIG_Vuln_IDs"] = stig_cols.apply(lambda x: x[0])
df["STIG_Severity"] = stig_cols.apply(lambda x: x[1])
df["STIG_CAT"] = stig_cols.apply(lambda x: x[2])
df["STIG_Titles"] = stig_cols.apply(lambda x: x[3])
df["STIG_Match_Method"] = stig_cols.apply(lambda x: x[4])
df["STIG_Match_Score"] = stig_cols.apply(lambda x: x[5])
df["STIG_Match_Notes"] = stig_cols.apply(lambda x: x[6])

# -----------------------------
# CIS mapping (from JSON generated from PDF)
# -----------------------------
df["CIS_Benchmark_Refs"] = df["Finding_Parts"].apply(format_cis_refs)

# -----------------------------
# Metrics Summary
# -----------------------------
def safe_count(mask):
    return int(mask.sum()) if hasattr(mask, "sum") else int(mask)

metrics = [
    ("Total rules reviewed", len(df)),
    ("Accept rules", safe_count(df["Action"].str.lower().str.strip() == "accept")),
    ("Deny/Drop rules", safe_count(df["Action"].str.lower().str.strip() != "accept")),
    ("Rules with findings (any type)", safe_count(df["Finding"].astype(str).str.strip() != "")),
    ("High risk rules", safe_count(df["Risk_Level"] == "High")),
    ("Medium risk rules", safe_count(df["Risk_Level"] == "Medium")),
    ("Low risk rules", safe_count(df["Risk_Level"] == "Low")),
    ("Overly broad (ANY src/dst)", safe_count(df["Flag_Any_SrcDst"] == "Yes")),
    ("Overly broad (large subnet)", safe_count(df["Flag_Broad_SrcDst"] == "Yes")),
    ("Unrestricted services (ANY)", safe_count(df["Flag_Any_Services"] == "Yes")),
    ("Allow rules without logging", safe_count(df["Flag_No_Logging"] == "Yes")),
    ("Unused rules (0 hits)", safe_count(df["Unused_Rule"] == "Yes")),
    ("Shadowed rules", safe_count(df["Flag_Shadowed"] == "Yes")),
    ("Redundant rules", safe_count(df["Flag_Redundant"] == "Yes")),
]
metrics_df = pd.DataFrame(metrics, columns=["Metric", "Count"])

# -----------------------------
# Final column ordering
# -----------------------------
core_cols = ["Rule_Number", "Rule_Name", "Source", "Destination", "Services", "Action", "Logging", "Install_On"]
analysis_cols = [
    "Hit_Count", "Last_Hit_Date", "Unused_Rule",
    "Flag_Any_SrcDst", "Flag_Broad_SrcDst", "Flag_Any_Services", "Flag_No_Logging",
    "Flag_Shadowed", "Shadowed_By_Rule", "Flag_Redundant",
    "Risk_Level",
    "Finding", "Remediation",
    "STIG_Vuln_IDs", "STIG_Severity", "STIG_CAT", "STIG_Titles",
    "STIG_Match_Method", "STIG_Match_Score", "STIG_Match_Notes",
    "CIS_Benchmark_Refs",
    "NIST_Control"
]
final_cols = [c for c in core_cols + analysis_cols if c in df.columns] + [c for c in df.columns if c not in set(core_cols + analysis_cols)]
df = df[final_cols]

# -----------------------------
# Excel Export + Formatting
# -----------------------------
if os.path.exists(output_file):
    os.remove(output_file)

FILL_RISK = {
    "High": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "Medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "Low": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
}
FILL_CAT = {
    "CAT I": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "CAT II": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "CAT III": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Firewall Rules & Findings", index=False)
    metrics_df.to_excel(writer, sheet_name="Metrics Summary", index=False)

    wb = writer.book
    ms = wb["Metrics Summary"]
    ms.freeze_panes = "A2"
    for cell in ms[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ms.column_dimensions["A"].width = 46
    ms.column_dimensions["B"].width = 12

    ws = wb["Firewall Rules & Findings"]
    ws.freeze_panes = "A2"
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(14, min(70, len(str(cell.value)) + 2))

    colnames = [c.value for c in ws[1]]
    risk_col = colnames.index("Risk_Level") + 1 if "Risk_Level" in colnames else None
    cat_col = colnames.index("STIG_CAT") + 1 if "STIG_CAT" in colnames else None

    for row in range(2, ws.max_row + 1):
        if risk_col:
            lvl = ws.cell(row=row, column=risk_col).value
            if lvl in FILL_RISK:
                ws.cell(row=row, column=risk_col).fill = FILL_RISK[lvl]
        if cat_col:
            cat = ws.cell(row=row, column=cat_col).value
            if cat in FILL_CAT:
                ws.cell(row=row, column=cat_col).fill = FILL_CAT[cat]

print(f"✔ v8 export complete: {output_file}")
